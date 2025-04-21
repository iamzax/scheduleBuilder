import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import copy
import os
from streamlit_sortables import sort_items

# --- Custom Colors ---
BACKGROUND_COLOR = "#FDF6F6"   # Background
BUTTON_COLOR = "#002C54"
TEXT_COLOR = "#000000"

EMPLOYEE_FILE = "employees.txt"

# --- Page Setup ---
st.set_page_config(page_title="📅 Schedule Builder", page_icon="📅", layout="centered")

# --- Custom Styling ---
st.markdown(
    f"""
    <style>
    .stApp {{
        background-color: {BACKGROUND_COLOR};
        color: {TEXT_COLOR};
    }}
    .stButton>button {{
        background-color: {BUTTON_COLOR};
        color: white;
        border-radius: 8px;
        height: 3em;
        width: 100%;
        font-size: 16px;
        font-weight: bold;
    }}
    .stDownloadButton>button {{
        background-color: {BUTTON_COLOR};
        color: white;
        border-radius: 8px;
        height: 3em;
        width: 100%;
        font-size: 16px;
        font-weight: bold;
    }}
    input, textarea {{
        background-color: #FBEAEB;
        color: white;
    }}
    .css-1cpxqw2, .css-1v0mbdj {{
        background-color: #FBEAEB;
        color: white;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

# --- Functions ---

def load_employees():
    if not os.path.exists(EMPLOYEE_FILE):
        return []
    with open(EMPLOYEE_FILE, "r") as f:
        employees = [line.strip() for line in f if line.strip()]
    return employees

def save_employees(employees):
    with open(EMPLOYEE_FILE, "w") as f:
        for emp in employees:
            f.write(emp + "\n")

def autofit_excel(workbook):
    ws = workbook.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_value = str(cell.value)
                if '\n' in cell_value:
                    lines = cell_value.split('\n')
                    longest_line = max(len(line) for line in lines)
                    max_length = max(max_length, longest_line)
                else:
                    max_length = max(max_length, len(cell_value))
        
        # --- SETTING MINIMUM COLUMN WIDTH TO 50
        adjusted_width = max(max_length + 2, 15)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows():
        for cell in row:
            alignment = copy.copy(cell.alignment)
            alignment.wrap_text = True
            cell.alignment = alignment
            cell.border = thin_border
        # No need to touch row height anymore!
    
    # Set narrow margins
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

# --- Load Employee Data ---
if "employees" not in st.session_state:
    st.session_state.employees = load_employees()

# --- App Title ---
st.markdown(f"<h1 style='text-align: center; color: {TEXT_COLOR};'>📅 Schedule Builder</h1>", unsafe_allow_html=True)

# --- Employee Manager ---
st.markdown(f"<h3 style='color: {TEXT_COLOR};'>Employee List:</h3>", unsafe_allow_html=True)

# Drag-and-Drop Sorting
new_order = sort_items(st.session_state.employees, direction="vertical")
if new_order != st.session_state.employees:
    st.session_state.employees = new_order
    save_employees(st.session_state.employees)
    st.success("Employee list reordered!")
    st.rerun()

# --- Add New Employee ---
st.markdown(f"<h4 style='color: {TEXT_COLOR};'>Add a New Employee:</h4>", unsafe_allow_html=True)
new_employee = st.text_input(label=" ", placeholder="Type name and press Enter")

if st.button("Add Employee"):
    if new_employee.strip():
        if new_employee.strip() not in st.session_state.employees:
            st.session_state.employees.append(new_employee.strip())
            save_employees(st.session_state.employees)
            st.success(f"{new_employee} added successfully.")
            st.rerun()
        else:
            st.warning("⚠️ This employee already exists!")
    else:
        st.warning("⚠️ Please enter a valid name.")

# --- Remove Employees ---
st.markdown(f"<h4 style='color: {TEXT_COLOR};'>Remove Employees:</h4>", unsafe_allow_html=True)

remove_selected = st.multiselect(label=" ", options=st.session_state.employees, key="remove_select_box")

if st.button("Remove Selected"):
    if remove_selected:
        # Remove the selected employees
        updated_list = [emp for emp in st.session_state.employees if emp not in remove_selected]
        st.session_state.employees = updated_list
        save_employees(updated_list)
        st.success("Selected employee(s) removed successfully.")
        st.rerun()
    else:
        st.warning("⚠️ No employee selected for removal.")

# --- Divider ---
st.markdown("---")

# --- Schedule Builder ---
st.markdown(f"<h3 style='color: {TEXT_COLOR};'>Select Starting Monday:</h3>", unsafe_allow_html=True)

selected_date = st.date_input("Pick a date", value=datetime.today())

if selected_date.weekday() != 0:
    st.warning("⚠️ Selected date is not a Monday. It will adjust automatically.")

if st.button("Generate Schedule"):
    if not st.session_state.employees:
        st.error("No employees to schedule! Please add employees first.")
    else:
        # Adjust to Monday
        if selected_date.weekday() != 0:
            selected_date -= timedelta(days=selected_date.weekday())

        days = [(selected_date + timedelta(days=i)).strftime("%Y-%m-%d") + "\n" + (selected_date + timedelta(days=i)).strftime("%A") for i in range(7)]

        df = pd.DataFrame(columns=["Employee"] + days)
        df["Employee"] = st.session_state.employees

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Schedule")

        output.seek(0)
        wb = load_workbook(filename=output)

        # Apply Formatting
        autofit_excel(wb)

        # Set Landscape Layout
        ws = wb.active
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER 

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.success("Schedule generated!")

        st.download_button(label="📥 Download Schedule",
                           data=final_output,
                           file_name="schedule.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")